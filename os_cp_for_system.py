# import subprocess
# import time
# import psutil
# from plyer import notification  # For showing desktop notifications
# import tkinter as tk
# import threading
# import openpyxl  # For working with Excel files

# # Initial Configuration
# MEMORY_LIMIT_MB = 500  # Set initial memory limit in MB
# CHECK_INTERVAL = 0.5  # Check every second
# warning_triggered = False  # Flag to track if the warning has already been triggered
# last_focused_window = None  # Store the last focused window to track changes
# EXCEL_FILE = "memory_limits.xlsx"  # Excel file to store app memory limits

# # Load or create the Excel file
# def load_or_create_excel():
#     try:
#         wb = openpyxl.load_workbook(EXCEL_FILE)
#         sheet = wb.active
#     except FileNotFoundError:
#         wb = openpyxl.Workbook()
#         sheet = wb.active
#         sheet.append(["App Name", "Memory Limit (MB)"])  # Add header row
#         wb.save(EXCEL_FILE)
#     return wb, sheet

# def get_focused_window_process():
#     """Get the process name of the currently focused window using xdotool."""
#     try:
#         # Get the active window ID using xdotool
#         window_id = subprocess.run(
#             ["xdotool", "getactivewindow"],
#             capture_output=True,
#             text=True,
#             check=True
#         ).stdout.strip()

#         if window_id:
#             # Get the PID of the active window using xdotool
#             pid = subprocess.run(
#                 ["xdotool", "getwindowpid", window_id],
#                 capture_output=True,
#                 text=True,
#                 check=True
#             ).stdout.strip()

#             if pid.isdigit():
#                 pid = int(pid)
#                 process = psutil.Process(pid)
#                 return process.name()
#         return None
#     except Exception as e:
#         print(f"Error in get_focused_window_process: {e}")
#         return None

# def show_notification(message):
#     """Show system-wide notification using plyer."""
#     notification.notify(
#         title="Memory Usage Alert",
#         message=message,
#         timeout=10  # Notification duration in seconds
#     )
#     print(f"Notification: {message}")

# def get_memory_limit_from_excel(sheet, app_name):
#     """Check the current memory limit for the app from Excel."""
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         if row[0].lower() == app_name.lower():
#             return row[1]
#     return None

# def update_memory_limit_in_excel(sheet, app_name, new_limit):
#     """Update the memory limit for the app in Excel."""
#     for row in sheet.iter_rows(min_row=2):
#         if row[0].value.lower() == app_name.lower():
#             row[1].value = new_limit
#             break
#     else:
#         sheet.append([app_name, new_limit])  # If app is not found, add a new row

# def monitor_memory(popup, label, wb, sheet):
#     """Monitor the memory usage of the focused window's application and system."""
#     global warning_triggered, last_focused_window

#     while True:
#         try:
#             # Get the process name of the focused window
#             app_name = get_focused_window_process()

#             # If the focused window has changed, reset the warning
#             if app_name != last_focused_window:
#                 last_focused_window = app_name
#                 warning_triggered = False  # Reset the warning trigger for the new focused window

#             if app_name is None:
#                 # Get system memory usage when no app is focused
#                 system_memory = psutil.virtual_memory()
#                 # used_system_memory_mb = system_memory.used / (1024 * 1024)  # Convert bytes to MB
#                 system_memory_percent = system_memory.percent
#                 label.config(
#                     text=f"No focused window\nSystem Used: {system_memory_percent:.1f} %"
#                 )
#                 time.sleep(CHECK_INTERVAL)
#                 continue

#             # Get all processes matching the app name
#             processes = [proc for proc in psutil.process_iter(['pid', 'name']) if app_name.lower() in proc.info['name'].lower()]

#             if not processes:
#                 label.config(text=f"No processes for {app_name}")
#                 time.sleep(CHECK_INTERVAL)
#                 continue

#             # Calculate total memory usage of the app
#             total_memory_usage_mb = sum(
#                 proc.memory_info().rss / (1024 * 1024) for proc in processes
#             )

#             # Get system memory usage
#             system_memory = psutil.virtual_memory()
#             system_memory_percent = system_memory.percent

#             # Update the label with memory usage
#             label.config(
#                 text=(
#                     f"{app_name}: {total_memory_usage_mb:.2f} MB\n"
#                     f"System Used: {system_memory_percent:.1f}%"
#                 )
#             )

#             # Get the memory limit for the app from the Excel sheet
#             memory_limit = get_memory_limit_from_excel(sheet, app_name)
#             if memory_limit is None:
#                 memory_limit = MEMORY_LIMIT_MB

#             # Check if memory usage exceeds the limit and the warning hasn't been triggered
#             if total_memory_usage_mb > memory_limit and not warning_triggered:
#                 message = f"Warning: {app_name} exceeded memory limit of {memory_limit} MB. Current memory usage: {total_memory_usage_mb:.2f} MB."
#                 show_notification(message)
#                 warning_triggered = True  # Set the flag so the warning is not triggered again

#                 # Open a new window to take new memory limit input
#                 open_memory_limit_input_window(app_name, memory_limit, wb, sheet)

#             time.sleep(CHECK_INTERVAL)

#         except psutil.NoSuchProcess:
#             time.sleep(CHECK_INTERVAL)
#         except Exception as e:
#             print(f"Error: {str(e)}")
#             break

# def open_memory_limit_input_window(app_name, current_limit, wb, sheet):
#     """Create a pop-up window to allow the user to input a new memory limit."""
#     input_window = tk.Toplevel()
#     input_window.geometry("300x150")
#     input_window.title(f"Set New Memory Limit for {app_name}")
#     input_window.attributes("-topmost", True)  # Keep the window always on top

#     label = tk.Label(input_window, text=f"Current Limit: {current_limit} MB\nEnter new memory limit (MB):", font=("Arial", 12))
#     label.pack(pady=10)

#     entry = tk.Entry(input_window, font=("Arial", 12))
#     entry.pack(pady=5)

#     def on_submit():
#         new_limit = entry.get()
#         if new_limit.isdigit():
#             new_limit = int(new_limit)
#             print(f"New memory limit set for {app_name}: {new_limit} MB")
#             update_memory_limit_in_excel(sheet, app_name, new_limit)
#             wb.save(EXCEL_FILE)  # Save the updated Excel file
#             input_window.destroy()  # Close the input window after submission
#         else:
#             print("Invalid input, please enter a valid number.")

#     submit_button = tk.Button(input_window, text="Submit", command=on_submit)
#     submit_button.pack(pady=10)

# def create_popup():
#     """Create a pop-up window to display memory usage."""
#     popup = tk.Toplevel()
#     popup.geometry("300x250")  # Adjust size of the pop-up to fit both app and system memory usage
#     popup.overrideredirect(True)  # Removes the title bar
#     popup.configure(bg="lightyellow")

#     # Get screen dimensions
#     screen_width = popup.winfo_screenwidth()
#     screen_height = popup.winfo_screenheight()

#     # Position the pop-up at the left bottom of the screen
#     x = 10  # 10 pixels from the left edge
#     y = screen_height - 270  # Offset from the bottom (150 height + 20 padding)
#     popup.geometry(f"300x250+{x}+{y}")

#     # Add a label to display the notification message
#     label = tk.Label(popup, text="Initializing...", font=("Arial", 12), bg="lightyellow", anchor="w")
#     label.pack(padx=10, pady=10)

#     # Load or create the Excel file
#     wb, sheet = load_or_create_excel()

#     # Start the memory monitoring in a separate thread
#     threading.Thread(target=monitor_memory, args=(popup, label, wb, sheet), daemon=True).start()

#     return popup

# if __name__ == "__main__":
#     # Main application window
#     root = tk.Tk()
#     root.withdraw()  # Hide the main window

#     # Create and display the pop-up
#     create_popup()

#     # Run the event loop
#     root.mainloop()



import subprocess
import time
import psutil
from plyer import notification  # For showing desktop notifications
import tkinter as tk
import threading
import openpyxl  # For working with Excel files

# Initial Configuration
MEMORY_LIMIT_MB = 500  # Set initial memory limit in MB
CHECK_INTERVAL = 0.5  # Check every half-second
warning_triggered = False  # Flag to track if the warning has already been triggered
last_focused_window = None  # Store the last focused window to track changes
EXCEL_FILE = "memory_limits.xlsx"  # Excel file to store app memory limits

# Load or create the Excel file
def load_or_create_excel():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["App Name", "Memory Limit (MB)"])  # Add header row
        wb.save(EXCEL_FILE)
    return wb, sheet

def get_focused_window_process():
    """Get the process name of the currently focused window using xdotool."""
    try:
        # Get the active window ID using xdotool
        window_id = subprocess.run(
            ["xdotool", "getactivewindow"],
            capture_output=True,
            text=True,
            check=True
        ).stdout.strip()

        if window_id:
            # Get the PID of the active window using xdotool
            pid = subprocess.run(
                ["xdotool", "getwindowpid", window_id],
                capture_output=True,
                text=True,
                check=True
            ).stdout.strip()

            if pid.isdigit():
                pid = int(pid)
                process = psutil.Process(pid)
                return process.name()
        return None
    except Exception as e:
        print(f"Error in get_focused_window_process: {e}")
        return None

def show_notification(message):
    """Show system-wide notification using plyer."""
    notification.notify(
        title="Memory Usage Alert",
        message=message,
        timeout=10  # Notification duration in seconds
    )
    print(f"Notification: {message}")

def get_memory_limit_from_excel(sheet, app_name):
    """Check the current memory limit for the app from Excel."""
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0].lower() == app_name.lower():
            return row[1]
    return None

def update_memory_limit_in_excel(sheet, app_name, new_limit):
    """Update the memory limit for the app in Excel."""
    for row in sheet.iter_rows(min_row=2):
        if row[0].value.lower() == app_name.lower():
            row[1].value = new_limit
            break
    else:
        sheet.append([app_name, new_limit])  # If app is not found, add a new row

def monitor_memory(label, wb, sheet):
    """Monitor the memory usage of the focused window's application and system."""
    global warning_triggered, last_focused_window

    while True:
        try:
            # Get the process name of the focused window
            app_name = get_focused_window_process()

            # If the focused window has changed, reset the warning flag
            if app_name != last_focused_window:
                last_focused_window = app_name
                warning_triggered = False  # Reset for new focused window

            if app_name is None:
                # When no app is focused, show system memory usage
                system_memory = psutil.virtual_memory()
                system_memory_percent = system_memory.percent
                label.config(
                    text=f"No focused window\nSystem Used: {system_memory_percent:.1f}%"
                )
                time.sleep(CHECK_INTERVAL)
                continue

            # Get all processes matching the app name
            processes = [proc for proc in psutil.process_iter(['pid', 'name']) if app_name.lower() in proc.info['name'].lower()]

            if not processes:
                label.config(text=f"No processes for {app_name}")
                time.sleep(CHECK_INTERVAL)
                continue

            # Calculate total memory usage of the app in MB
            total_memory_usage_mb = sum(
                proc.memory_info().rss / (1024 * 1024) for proc in processes
            )

            # Get system memory usage
            system_memory = psutil.virtual_memory()
            system_memory_percent = system_memory.percent

            # Update the label with current usage
            label.config(
                text=(
                    f"{app_name}:\n{total_memory_usage_mb:.2f} MB\n"
                    f"System Used: {system_memory_percent:.1f}%"
                )
            )

            # Get the memory limit for the app from the Excel sheet
            memory_limit = get_memory_limit_from_excel(sheet, app_name)
            if memory_limit is None:
                memory_limit = MEMORY_LIMIT_MB

            # Check if memory usage exceeds the limit and if the warning hasn’t been triggered yet
            if total_memory_usage_mb > memory_limit and not warning_triggered:
                message = (f"Warning: {app_name} exceeded memory limit of {memory_limit} MB.\n"
                           f"Current usage: {total_memory_usage_mb:.2f} MB.")
                show_notification(message)
                warning_triggered = True  # Prevent repeated warnings

                # Open a new window to accept a new memory limit
                open_memory_limit_input_window(app_name, memory_limit, wb, sheet)

            time.sleep(CHECK_INTERVAL)

        except psutil.NoSuchProcess:
            time.sleep(CHECK_INTERVAL)
        except Exception as e:
            print(f"Error in monitor_memory: {e}")
            break

def open_memory_limit_input_window(app_name, current_limit, wb, sheet):
    """Open a pop-up window to allow the user to input a new memory limit."""
    input_window = tk.Toplevel()
    input_window.geometry("300x150")
    input_window.title(f"Set New Memory Limit for {app_name}")
    input_window.attributes("-topmost", True)  # Ensure the window stays on top

    label = tk.Label(
        input_window, 
        text=f"Current Limit: {current_limit} MB\nEnter new memory limit (MB):", 
        font=("Segoe UI", 12)
    )
    label.pack(pady=10)

    entry = tk.Entry(input_window, font=("Segoe UI", 12))
    entry.pack(pady=5)

    def on_submit():
        new_limit = entry.get()
        if new_limit.isdigit():
            new_limit_int = int(new_limit)
            print(f"New memory limit set for {app_name}: {new_limit_int} MB")
            update_memory_limit_in_excel(sheet, app_name, new_limit_int)
            wb.save(EXCEL_FILE)  # Save changes
            input_window.destroy()  # Close the input window
        else:
            print("Invalid input, please enter a valid number.")

    submit_button = tk.Button(input_window, text="Submit", command=on_submit, font=("Segoe UI", 12))
    submit_button.pack(pady=10)

def create_popup(wb, sheet):
    """Create an aesthetically improved pop-up window to display memory usage."""
    popup = tk.Toplevel()
    popup.geometry("320x120")  # Set window size
    popup.overrideredirect(True)  # Remove title bar for a modern look
    popup.configure(bg="#2E3440")  # Use a dark background (Nord theme)

    # Get screen dimensions and position the pop-up at the bottom left
    screen_width = popup.winfo_screenwidth()
    screen_height = popup.winfo_screenheight()
    x = 20  # 20 pixels from the left edge
    y = screen_height - 150  # Adjust from the bottom
    popup.geometry(f"320x120+{x}+{y}")

    # Add a label to display the memory usage text
    label = tk.Label(
        popup,
        text="Initializing...",
        font=("Segoe UI", 12, "bold"),
        fg="#D8DEE9",  # Light text color
        bg="#2E3440",  # Match background
        anchor="w"
    )
    label.pack(padx=15, pady=15, fill='both', expand=True)

    # Add a custom close button (×)
    close_button = tk.Button(
        popup,
        text="×",
        font=("Segoe UI", 14, "bold"),
        fg="white",
        bg="#BF616A",
        command=popup.destroy,
        bd=0,
        relief="flat",
        padx=5,
        pady=2
    )
    close_button.place(relx=0.9, rely=0.05)

    # Start the memory monitoring thread that will update the label
    threading.Thread(target=monitor_memory, args=(label, wb, sheet), daemon=True).start()

    return popup

if __name__ == "__main__":
    # Main application window (hidden)
    root = tk.Tk()
    root.withdraw()

    # Load or create the Excel file
    wb, sheet = load_or_create_excel()

    # Create and display the aesthetic pop-up
    create_popup(wb, sheet)

    # Start the Tkinter event loop
    root.mainloop()
