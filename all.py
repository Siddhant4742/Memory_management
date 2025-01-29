import psutil
import time
import tkinter as tk
from tkinter import ttk
import openpyxl
import threading
class SystemMonitor:
    def __init__(self, root):
        self.root = root
        self.root.title("System Monitor")

        # Create a Treeview to display process information
        columns = ("PID", "Name", "CPU%", "Memory (MB)")
        self.tree = ttk.Treeview(root, columns=columns, show="headings")
        self.tree.heading("PID", text="PID")
        self.tree.heading("Name", text="Name")
        self.tree.heading("CPU%", text="CPU %")
        self.tree.heading("Memory (MB)", text="Memory (MB)")
        self.tree.pack(fill="both", expand=True)

        # Create a label to display system-wide memory usage
        self.system_memory_label = tk.Label(root, text="")
        self.system_memory_label.pack()

        # Load or create the Excel file
        self.filename = "process_data.xlsx"
        self.sheet_name = "Sheet1"
        self.wb, self.sheet = self.load_or_create_excel()

        # Start the monitoring in a separate thread
        self.update_thread = None
        self.start_monitoring()

    def load_or_create_excel(self):
        try:
            wb = openpyxl.load_workbook(self.filename)
            sheet = wb[self.sheet_name]
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = self.sheet_name
            sheet.append(["PID", "Name", "CPU%", "Memory (MB)"])  # Add header row
            wb.save(self.filename)
        return wb, sheet

    def start_monitoring(self):
        self.update_thread = threading.Thread(target=self.update_data)
        self.update_thread.daemon = True
        self.update_thread.start()

    def update_data(self):
        while True:
            try:
                # Get system memory usage
                system_memory = psutil.virtual_memory()
                used_system_memory_mb = system_memory.used / (1024 * 1024)
                self.system_memory_label.config(
                    text=f"System Memory Used: {used_system_memory_mb:.2f} MB"
                )

                # Clear existing data in the Treeview
                for i in self.tree.get_children():
                    self.tree.delete(i)

                # Get all running processes
                for proc in psutil.process_iter(['pid', 'name', 'cpu_percent', 'memory_info']):
                    try:
                        cpu_usage = proc.info['cpu_percent']
                        memory_usage_mb = proc.info['memory_info'].rss / (1024 * 1024)
                        self.tree.insert("", "end", values=(proc.info['pid'], proc.info['name'], cpu_usage, f"{memory_usage_mb:.2f}"))

                        # Save data to Excel (optional)
                        self.save_process_data_to_excel(proc.info['pid'], proc.info['name'], cpu_usage, memory_usage_mb)

                    except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                        pass

                time.sleep(1)  # Update every second

            except Exception as e:
                print(f"Error: {str(e)}")
                break

    def save_process_data_to_excel(self, pid, name, cpu_usage, memory_usage_mb):
        # Check if data for this PID already exists
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == pid:
                # Update existing row
                row[1] = name
                row[2] = cpu_usage
                row[3] = memory_usage_mb
                return

        # Add new row to Excel
        self.sheet.append([pid, name, cpu_usage, memory_usage_mb])
        self.wb.save(self.filename)

if __name__ == "__main__":
    root = tk.Tk()
    app = SystemMonitor(root)
    root.mainloop()